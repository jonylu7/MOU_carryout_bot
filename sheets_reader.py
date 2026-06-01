"""
carry_out_automation/sheets_reader.py

從 Google Drive 下載 .xlsx 申請表，讀取待處理的攜出申請，
並在完成後更新「送出申請」欄位後上傳回 Drive。

支援兩種檔案類型：
  - Google Sheets（spreadsheets/d/…）：export 下載、Sheets API 寫回
  - Drive 上的 .xlsx（file/d/…）：get_media 下載、files.update 上傳

依賴：
    pip install google-auth google-auth-oauthlib google-api-python-client openpyxl pandas
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Optional

import pandas as pd
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaIoBaseUpload

import config

logger = logging.getLogger(__name__)

SCOPES = [
    "https://www.googleapis.com/auth/drive",               # 讀寫 Drive 檔案
    "https://www.googleapis.com/auth/spreadsheets",        # Sheets API 寫回
]

GSHEET_MIMETYPE = "application/vnd.google-apps.spreadsheet"

# 模組級快取，避免同次執行多次呼叫 API 查詢 mimeType
_file_mimetype_cache: dict[str, str] = {}

# 試算表結構：第 0~1 列為公告/說明，第 2 列為欄位標題，第 3 列起為資料
# header=2 表示以第 2 列（0-indexed）作為欄位標題
SHEET_HEADER_ROW   = 2          # 0-indexed，傳給 pd.read_excel header=
# df row 0 對應試算表第幾列（1-indexed，Sheets API 用）
SHEET_DATA_ROW_START = SHEET_HEADER_ROW + 2   # = 4（標題列 3 + 資料從列 4 起）


# ─────────────────────────────────────────────────
# Google API 服務初始化
# ─────────────────────────────────────────────────

def _get_creds() -> Credentials:
    return Credentials.from_service_account_file(
        str(config.GOOGLE_SERVICE_ACCOUNT_KEY),
        scopes=SCOPES,
    )

def _get_drive_service():
    return build("drive", "v3", credentials=_get_creds())

def _get_sheets_service():
    return build("sheets", "v4", credentials=_get_creds())


def _normalize_col(name: str) -> str:
    """
    把試算表欄名中的換行去掉，並截去括號/逗號後的說明文字，
    讓欄名可以用簡短的鍵值查詢。

    例：
        '預定攜出\\n日期'                → '預定攜出日期'
        '所在電腦\\n(請將檔案放在桌面)' → '所在電腦'
        '範例連結，\\n若有多個...'       → '範例連結'
    """
    s = re.sub(r"\n", "", str(name)).strip()       # 去除所有換行
    s = re.split(r"[（(（，,]", s)[0].strip()  # 截去第一個括號或逗號之後
    return s


def _is_google_sheet(file_id: str) -> bool:
    """回傳 True 若 file_id 指向 Google Sheets（而非 Drive 上的 .xlsx）。"""
    if file_id not in _file_mimetype_cache:
        service = _get_drive_service()
        meta = service.files().get(fileId=file_id, fields="mimeType").execute()
        _file_mimetype_cache[file_id] = meta.get("mimeType", "")
    return _file_mimetype_cache[file_id] == GSHEET_MIMETYPE


# ─────────────────────────────────────────────────
# 資料結構
# ─────────────────────────────────────────────────

@dataclass
class CarryOutRequest:
    """代表 xlsx 中一列攜出申請"""
    row_index: int          # DataFrame 的 index（0-based）
    填寫日期: str
    預定攜出日期: str
    預定攜出時間: str
    填寫人: str
    填寫人信箱: str
    所在電腦: str
    議題範圍: str
    已填寫攜出範例: bool
    承辦人: str
    親自到場: bool
    檔案名稱: str
    檔案屬性: str
    格式內容說明: str
    範例連結: str
    特殊備註: str
    送出申請: bool

    @property
    def carry_out_date(self) -> Optional[date]:
        for fmt in (
            "%Y/%m/%d",
            "%Y-%m-%d",
            "%Y-%m-%d %H:%M:%S",   # Google Sheets export 可能帶時間
            "%m/%d/%Y",
            "%Y年%m月%d日",
        ):
            try:
                return datetime.strptime(str(self.預定攜出日期).strip(), fmt).date()
            except ValueError:
                continue
        # 若已是 datetime/date 物件（pandas 解析過的）
        if isinstance(self.預定攜出日期, (datetime, date)):
            d = self.預定攜出日期
            return d.date() if isinstance(d, datetime) else d
        return None


# ─────────────────────────────────────────────────
# 下載 xlsx
# ─────────────────────────────────────────────────

def download_xlsx(file_id: str) -> bytes:
    """
    從 Google Drive 下載指定 file_id 的內容，回傳 xlsx bytes。

    - Google Sheets（spreadsheets/d/…）：使用 export API 轉成 xlsx
    - 一般 Drive .xlsx 檔（file/d/…）：直接 get_media
    """
    service = _get_drive_service()
    if _is_google_sheet(file_id):
        logger.info(f"偵測到 Google Sheets，使用 export 下載…")
        request = service.files().export(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = service.files().get_media(fileId=file_id)

    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    logger.info(f"已下載 Drive 檔案 {file_id}（{buf.tell()} bytes）")
    return buf.getvalue()


# ─────────────────────────────────────────────────
# 讀取申請
# ─────────────────────────────────────────────────

def fetch_pending_requests(
    days_ahead: int = config.DAYS_AHEAD_FILTER,
) -> tuple[list[CarryOutRequest], pd.DataFrame, bytes]:
    """
    從 Google Drive 下載 xlsx，回傳：
      - list[CarryOutRequest]：待處理申請
      - df：完整 DataFrame（稍後寫回用）
      - raw_bytes：原始 xlsx bytes（備用）
    """
    raw = download_xlsx(config.DRIVE_FILE_ID)
    df = pd.read_excel(
        io.BytesIO(raw),
        sheet_name=config.SHEET_NAME,
        dtype=str,                      # 全部先讀成字串，避免日期被亂轉
        header=SHEET_HEADER_ROW,        # 跳過前 2 列公告/說明文字
    )

    # 標準化欄名：去除換行、截去括號/逗號後說明
    df.columns = [_normalize_col(c) for c in df.columns]

    cutoff = date.today() + timedelta(days=days_ahead)
    pending: list[CarryOutRequest] = []

    for idx, row in df.iterrows():
        # 遇到「已完成攜出」分隔列就停止，其下方皆為已完成申請
        first_cell = str(list(row)[0])
        if "已完成攜出" in first_cell:
            logger.info(f"遇到分隔列（{first_cell[:20]}），停止讀取")
            break

        req = CarryOutRequest(
            row_index=int(idx),
            填寫日期=row.get("填寫日期", ""),
            預定攜出日期=row.get("預定攜出日期", ""),
            預定攜出時間=row.get("預定攜出時間", ""),
            填寫人=row.get("填寫人", ""),
            填寫人信箱=row.get("填寫人信箱", ""),
            所在電腦=row.get("所在電腦", ""),
            議題範圍=row.get("攜出內容所屬議題範圍", ""),
            已填寫攜出範例=str(row.get("是否已填寫攜出範例", "")).upper() in ("TRUE", "是", "✓"),
            承辦人=row.get("承辦人", ""),
            親自到場=str(row.get("是否親自到場進行說明", "")).upper() in ("TRUE", "是", "✓"),
            檔案名稱=row.get("攜出檔案名稱", ""),
            檔案屬性=row.get("攜出檔案屬性", ""),
            格式內容說明=row.get("攜出檔案格式內容說明", ""),
            範例連結=row.get("範例連結", ""),
            特殊備註=row.get("特殊備註", ""),
            送出申請=False,
        )

        carry_date = req.carry_out_date
        if carry_date is None:
            logger.warning(f"第 {idx} 列日期無法解析：「{req.預定攜出日期}」，略過")
            continue

        pending.append(req)
        logger.info(f"待處理：{req.填寫人}  攜出日：{req.預定攜出日期}")

    return pending, df, raw


def group_by_date_and_topic(
    requests: list[CarryOutRequest],
) -> dict[tuple[str, str], list[CarryOutRequest]]:
    """依 (預定攜出日期, 議題範圍) 分組，同一天同議題合成一份攜出單。"""
    groups: dict[tuple, list[CarryOutRequest]] = {}
    for req in requests:
        key = (req.預定攜出日期, req.議題範圍)
        groups.setdefault(key, []).append(req)
    return groups


# ─────────────────────────────────────────────────
# 寫回（更新「送出申請」欄）
# ─────────────────────────────────────────────────

def _col_letter(n: int) -> str:
    """將 1-based 欄號轉為 A1 欄位字母（1→A, 27→AA）。"""
    result = ""
    while n:
        n, r = divmod(n - 1, 26)
        result = chr(65 + r) + result
    return result


def mark_as_submitted(
    row_indices: list[int],
    df: pd.DataFrame,
) -> None:
    """
    將指定列的「送出申請」設為 TRUE。

    - Google Sheets：用 Sheets API batchUpdate 只改動目標儲存格（不破壞其他格式）
    - Drive .xlsx：序列化後整檔上傳覆蓋
    """
    # 先更新記憶體中的 df（供後續邏輯使用）
    for idx in row_indices:
        df.at[idx, "送出申請"] = "TRUE"

    if _is_google_sheet(config.DRIVE_FILE_ID):
        _mark_submitted_via_sheets_api(row_indices, df)
    else:
        _mark_submitted_via_drive_upload(df)


def _mark_submitted_via_sheets_api(row_indices: list[int], df: pd.DataFrame) -> None:
    """Sheets API：只更新「送出申請」欄的指定儲存格。"""
    if "送出申請" not in df.columns:
        logger.error("找不到「送出申請」欄位，無法更新")
        return

    col_idx = df.columns.tolist().index("送出申請")
    col_letter = _col_letter(col_idx + 1)   # 1-based → 欄位字母

    # DataFrame 0-based index → Sheet 列號（1-indexed）
    # 試算表前 3 列為公告+標題，資料從第 SHEET_DATA_ROW_START 列起
    data = [
        {
            "range": f"{config.SHEET_NAME}!{col_letter}{idx + SHEET_DATA_ROW_START}",
            "values": [["TRUE"]],
        }
        for idx in row_indices
    ]

    sheets = _get_sheets_service()
    sheets.spreadsheets().values().batchUpdate(
        spreadsheetId=config.DRIVE_FILE_ID,
        body={"valueInputOption": "RAW", "data": data},
    ).execute()
    logger.info(f"已透過 Sheets API 將 {len(row_indices)} 列標記為已送出")


def _ensure_google_sheet(file_id: str) -> tuple[str, bool]:
    """
    若 file_id 是原生 Google Sheets 則直接回傳；
    若是 Drive 上的 xlsx 則複製一份並轉為 Sheets，回傳 (copy_id, True)。
    呼叫方有責任在使用完畢後刪除副本（第二個回傳值 = True 時需刪除）。
    """
    if _is_google_sheet(file_id):
        return file_id, False
    # xlsx on Drive → 複製並轉換為 Google Sheets
    service = _get_drive_service()
    logger.info(f"偵測到 Drive xlsx，複製並轉為 Google Sheets：{file_id}")
    copy = service.files().copy(
        fileId=file_id,
        body={"mimeType": GSHEET_MIMETYPE},
    ).execute()
    return copy["id"], True


def _delete_file(file_id: str) -> None:
    """刪除 Drive 上的檔案（用於清理暫時副本）。"""
    try:
        _get_drive_service().files().delete(fileId=file_id).execute()
        logger.info(f"已刪除暫時副本：{file_id}")
    except Exception as e:
        logger.warning(f"刪除副本失敗（{file_id}）：{e}")


def _render_xlsx_range_as_png(
    xlsx_bytes: bytes,
    range_notation: str | None = None,
    sheet_name: str | None = None,
) -> bytes:
    """
    用 openpyxl + matplotlib 本地渲染 xlsx 表格為 PNG，不需要 Drive 副本或 Sheets API。
    sheet_name 為 None 時使用 active sheet；range_notation 為 None 時渲染整張工作表。
    """
    import openpyxl
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    plt.rcParams["font.family"] = [
        "DFKai-SB", "Kaiti SC", "Arial Unicode MS", "Heiti TC", "LiHei Pro",
        "Apple LiGothic", "DejaVu Sans"
    ]

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    if range_notation:
        raw = ws[range_notation]
        # normalize to list-of-lists
        if hasattr(raw, "value"):          # single Cell
            cell_rows = [[raw]]
        elif raw and not isinstance(raw[0], tuple):  # single row
            cell_rows = [list(raw)]
        else:
            cell_rows = [list(r) for r in raw]
    else:
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        cell_rows = [
            list(r) for r in ws.iter_rows(
                min_row=1, max_row=max_row, min_col=1, max_col=max_col
            )
        ]

    data = []
    for row in cell_rows:
        row_data = [str(c.value) if c.value is not None else "" for c in row]
        if any(row_data):   # 跳過完全空白列
            data.append(row_data)

    if not data:
        logger.warning(f"xlsx 範圍內無資料（range={range_notation}），回傳空 bytes")
        return b""

    n_cols = max(len(r) for r in data)
    data = [r + [""] * (n_cols - len(r)) for r in data]

    fig_w = max(6.0, min(20.0, n_cols * 2.0))
    fig_h = max(2.0, min(14.0, len(data) * 0.5))

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis("off")

    tbl = ax.table(cellText=data, bbox=[0, 0, 1, 1], cellLoc="left")
    tbl.auto_set_font_size(True)
    tbl.auto_set_column_width(list(range(n_cols)))

    for j in range(n_cols):
        tbl[0, j].set_facecolor("#4472C4")
        tbl[0, j].set_text_props(color="white", fontweight="bold")
    for i in range(1, len(data)):
        for j in range(n_cols):
            if i % 2 == 0:
                tbl[i, j].set_facecolor("#D9E1F2")

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", pad_inches=0.05)
    plt.close(fig)
    png_bytes = buf.getvalue()
    logger.info(
        f"matplotlib 本地渲染 PNG：{len(png_bytes)} bytes"
        f"（range={range_notation or '全頁'}）"
    )
    return png_bytes


def export_google_sheet_as_pdf(file_id: str) -> bytes:
    """
    將指定 Google Drive 檔案以 PDF 格式匯出，回傳 PDF bytes。
    若為 Drive 上的 xlsx（非原生 Sheets），先複製轉換再匯出，完成後刪除副本。
    """
    sheet_id, is_copy = _ensure_google_sheet(file_id)
    try:
        service = _get_drive_service()
        request = service.files().export(
            fileId=sheet_id,
            mimeType="application/pdf",
        )
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        logger.info(f"已匯出 PDF：{file_id}（{buf.tell()} bytes）")
        return buf.getvalue()
    finally:
        if is_copy:
            _delete_file(sheet_id)


def read_sheet_values(file_id: str) -> tuple[list[list], str]:
    """
    讀取試算表第一個工作表的所有儲存格值。
    - Google Sheets：透過 Sheets API 讀取
    - Drive xlsx：直接下載並用 openpyxl 解析（不建立 Drive 副本）
    回傳 (values: list[list[str]], sheet_title: str)。
    """
    if not _is_google_sheet(file_id):
        import openpyxl
        xlsx_bytes = download_xlsx(file_id)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
        ws = wb.active
        sheet_title = ws.title
        values = [
            [str(c) if c is not None else "" for c in row]
            for row in ws.iter_rows(values_only=True)
        ]
        logger.info(f"已本地解析 xlsx 儲存格值：{file_id}，{len(values)} 列")
        return values, sheet_title

    service = _get_sheets_service()
    meta = service.spreadsheets().get(
        spreadsheetId=file_id,
        fields="sheets.properties",
    ).execute()
    first_sheet = meta["sheets"][0]["properties"]
    sheet_title = first_sheet["title"]

    result = service.spreadsheets().values().get(
        spreadsheetId=file_id,
        range=sheet_title,
    ).execute()
    values = result.get("values", [])
    logger.info(f"已讀取試算表值：{file_id}，{len(values)} 列")
    return values, sheet_title


def export_sheet_range_as_png(file_id: str, range_notation: str | None = None) -> bytes:
    """
    匯出指定 range（或整頁）為 PNG bytes。
    下載為 xlsx（Google Sheets 用 files().export()，Drive xlsx 用 get_media()），
    再用 openpyxl + matplotlib 本地渲染，不依賴 Sheets PNG 匯出 URL（該 URL 對
    service account 不穩定，常回傳 400）。
    """
    xlsx_bytes = download_xlsx(file_id)
    return _render_xlsx_range_as_png(xlsx_bytes, range_notation)


def export_workbook_as_pngs(file_id: str) -> list[bytes]:
    """
    下載 workbook（Google Sheets 或 Drive xlsx），對每個非空 worksheet 渲染一張 PNG。
    回傳 list[bytes]，順序對應工作表順序。
    """
    import openpyxl

    xlsx_bytes = download_xlsx(file_id)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)

    images: list[bytes] = []
    for sname in wb.sheetnames:
        ws = wb[sname]
        if not ws.max_row or not ws.max_column:
            continue
        # 快速判斷是否有資料（掃前 10×10 格）
        has_data = any(
            ws.cell(r, c).value is not None
            for r in range(1, min(ws.max_row + 1, 11))
            for c in range(1, min(ws.max_column + 1, 11))
        )
        if not has_data:
            continue
        png = _render_xlsx_range_as_png(xlsx_bytes, None, sname)
        if png:
            images.append(png)
            logger.info(f"  工作表 '{sname}'：{len(png)} bytes")

    logger.info(f"workbook {file_id} 共渲染 {len(images)} 個工作表")
    return images


def _mark_submitted_via_drive_upload(df: pd.DataFrame) -> None:
    """Drive 上傳：序列化整份 xlsx 後覆蓋原檔（適用 Drive .xlsx）。"""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=config.SHEET_NAME, index=False)
    buf.seek(0)

    service = _get_drive_service()
    media = MediaIoBaseUpload(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=False,
    )
    service.files().update(fileId=config.DRIVE_FILE_ID, media_body=media).execute()
    logger.info(f"已將更新後的 xlsx 上傳回 Drive")
